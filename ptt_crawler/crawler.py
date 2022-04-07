'''
==========================================
Load packages
==========================================
'''
import requests        
import bs4            
import pandas as pd    
from glob import glob  

'''
==========================================
Crawler
==========================================
'''
query = 'enter_your_query'

for j in range(1,17):
    URL = "https://www.ptt.cc/bbs/Gossiping/search?page="+ str(j) + "&q=" + query    
    
    my_headers = {'cookie': 'over18=1;'}
    
    response = requests.get(URL, headers = my_headers)
           
    soup = bs4.BeautifulSoup(response.text,"html.parser")

    titles = soup.find_all('div','title')

    for t in titles:
        print(t.text)
    
    title_list = []
    
    for t in titles:
        title_list.append(t.text)
       
    df = pd.DataFrame()
    df ['title'] = title_list
    df.to_csv(str(j)+'.csv',index=False,encoding="utf-8-sig")
    
'''
==========================================
Combine csv files
==========================================
'''

files = glob('../Crawler/*.csv')

df = pd.concat((pd.read_csv(file, usecols=['title'], encoding='utf-8-sig') for file in files), ignore_index=True)

df.to_csv('datas.csv',index=False,encoding='utf-8-sig')

'''
==========================================
Through Jieba, find top 3 keywords.
==========================================
'''

import jieba
import csv
import operator
from collections import Counter

jieba.set_dictionary('../Crawler/dict.txt')
jieba.load_userdict('../Crawler/my.dict')

with open('sugar_apple.txt') as f:
    stops = f.read().split('\n')
    
with open('datas.csv', encoding='utf-8-sig') as csvfile:
    rows = csv.reader(csvfile)
    for row in rows:
        row = str(row)
        corpus = [] #語料庫
        
        corpus += [t for t in jieba.cut(row, cut_all=False) if t not in stops] 
        dic = {}
        for ele in corpus:
            if ele not in dic:
                dic[ele] = 1
            else:
                dic[ele] = dic[ele] + 1

        sorted_word = sorted(dic.items(), key=operator.itemgetter(1), reverse=True)
        print(sorted_word)
        with open('te.csv', 'a', encoding='utf-8-sig') as wf2:
            for ele in sorted_word:
                
                if len(ele[0]) >= 2:
                    world = ele[0] + ','
                    wf2.write(world)

te = open('te.csv', encoding='utf-8-sig').read()

list1 = te.split(',')

dir1 = Counter(list1)

with open('te1.csv', encoding='utf-8-sig', mode='w') as fp:
    fp.write('keyword,value,\n')
    for tag, count in dir1.items():
        fp.write('{},{}\n'.format(tag, count))


df = pd.read_csv('te1.csv', index_col="value")

df_top3 = df.sort_index(axis=0, ascending=False)[:10]

# save file
df_top3.to_csv('Keyword_top10.csv', encoding='utf-8-sig')

'''
==========================================
Selenium
==========================================
'''

from selenium.webdriver import Chrome
import time

browser = Chrome("../Crawler/chromedriver.exe")

browser.get("https://www.google.com")

inputElement = browser.find_element_by_name("q")

inputElement.send_keys("enter_keywords")

inputElement.submit()

time.sleep(3)

'''
==========================================
Put the URL into csv
==========================================
'''

from bs4 import BeautifulSoup

soup = BeautifulSoup(browser.page_source, 'html.parser')

path_list = []
nList = []

#Read five pages
for i in range(5): 
    
    time.sleep(3)  #buffer
    
    find_path = soup.find_all('div','tF2Cxc')
    name_datas=soup.find_all('h3',class_='LC20lb DKV0Md')
    
    for j in find_path:    
        path = j.find('a').get('href')
        path_list.append(path)
        #print(path)     
        
    for value in name_datas:  
        #print(value.text)
        nList.append(value.text)

    browser.find_element_by_link_text('下一頁').click()
        
df_html = pd.DataFrame({'title': nList,'URL': path_list})

#save
df_html.to_excel('keyword_result.xlsx',index = False , encoding = 'utf-8-sig')

browser.close()

'''
==========================================
beautify 
==========================================
'''
import openpyxl
from openpyxl.styles import Font,PatternFill 

fn = 'keyword_result.xlsx'
wb = openpyxl.load_workbook(fn)
ws = wb.active

print(wb.sheetnames)
# ['Sheet1']

ws.title = 'Search titles and URLs'

sheet = wb['Search titles and URLs']
title = sheet[1]  

#Change title font
font = Font(name='Courier', size=20, bold = True,color='000000') 

for row in ws.iter_rows(min_row=1, max_col=2, max_row=1):
    for cell in row:
        cell.font = font   
        print(cell)

#Change font size
for row in ws.iter_rows(min_row=2, max_col=2, max_row=(len(nList)+1)):
    for cell in row:
        cell.font = Font(name='Courier', size=12, color='000000')       
        print(cell)
        
#Change title color
fill_pattern = PatternFill(patternType='solid',fgColor='dcae96')

for row in ws.iter_rows(min_row=1, max_col=2, max_row=1):
    for cell in row:
        cell.fill = fill_pattern      
        print(cell)
        
wb.save(fn)  #save







